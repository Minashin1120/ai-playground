# Validates a spreadsheet cell address such as "B5" or "AD15".
_XLSX_CELL_RE = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]{0,6}$")


def _normalize_xlsx_color(color):
    """Normalize a user/model-provided color to openpyxl ARGB hex (FFRRGGBB).

    Accepts ``#RRGGBB`` / ``RRGGBB`` / 3-digit CSS shorthand / 8-digit ARGB.
    Returns ``None`` when the value cannot be parsed.
    """
    if color is None:
        return None
    text = str(color).strip()
    if not text:
        return None
    if text.startswith("#"):
        text = text[1:]
    text = text.strip()
    if len(text) == 3:
        text = "".join(ch * 2 for ch in text)
    if len(text) == 6:
        # openpyxl would treat a 6-digit value as ARGB with alpha 00
        # (fully transparent), so make 6-digit colors opaque.
        text = "FF" + text
    if len(text) != 8 or not re.fullmatch(r"[0-9A-Fa-f]{8}", text):
        return None
    return text.upper()


def _apply_cell_style(cell, style, label):
    """Apply a style dict (fill / font / border / alignment / numberFormat) to an openpyxl cell.

    Only the properties explicitly present in ``style`` are changed; every other
    formatting property on the cell is left untouched, so a style edit adds new
    formatting without destroying existing formatting.  Returns a list of
    human-readable error messages for invalid values (the valid parts are still
    applied).
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    errors = []
    if not isinstance(style, dict):
        return [f"{label}: style はオブジェクトで指定してください。"]

    fill = style.get("fill")
    if fill is not None:
        if not isinstance(fill, dict):
            errors.append(f"{label}: fill はオブジェクトで指定してください。")
        else:
            try:
                fill_type = str(fill.get("fillType") or "").strip().lower()
                color_raw = fill.get("color")
                if fill_type == "none":
                    cell.fill = PatternFill()
                elif color_raw is not None:
                    color = _normalize_xlsx_color(color_raw)
                    if color is None:
                        errors.append(f"{label}: 塗りつぶし色が不正です: {color_raw}")
                    else:
                        cell.fill = PatternFill(
                            start_color=color,
                            end_color=color,
                            fill_type=fill_type if fill_type in ("solid", "medium", "dark") else "solid",
                        )
                elif fill_type == "solid":
                    errors.append(f"{label}: 塗りつぶし色 (fill.color) を指定してください。")
            except Exception as exc:
                errors.append(f"{label}: 塗りつぶしの設定に失敗しました: {str(exc)[:80]}")

    font = style.get("font")
    if font is not None:
        if not isinstance(font, dict):
            errors.append(f"{label}: font はオブジェクトで指定してください。")
        else:
            try:
                cur = cell.font
                kw = {
                    "name": cur.name,
                    "size": cur.size,
                    "bold": cur.bold,
                    "italic": cur.italic,
                    "underline": cur.underline,
                    "strikethrough": cur.strikethrough,
                    "color": cur.color,
                    "vertAlign": cur.vertAlign,
                }
                if font.get("name") is not None:
                    kw["name"] = str(font["name"])
                if font.get("size") is not None:
                    try:
                        kw["size"] = float(font["size"])
                    except (TypeError, ValueError):
                        errors.append(f"{label}: フォントサイズが不正です: {font['size']}")
                if font.get("bold") is not None:
                    kw["bold"] = bool(font["bold"])
                if font.get("italic") is not None:
                    kw["italic"] = bool(font["italic"])
                if font.get("strikethrough") is not None:
                    kw["strikethrough"] = bool(font["strikethrough"])
                if font.get("underline") is not None:
                    u = str(font["underline"]).strip().lower()
                    if u in ("none", "single", "double", "singleaccounting", "doubleaccounting"):
                        kw["underline"] = u
                    else:
                        errors.append(f"{label}: underline の値が不正です: {font['underline']}")
                if font.get("color") is not None:
                    color = _normalize_xlsx_color(font["color"])
                    if color is None:
                        errors.append(f"{label}: 文字色が不正です: {font['color']}")
                    else:
                        kw["color"] = color
                cell.font = Font(**kw)
            except Exception as exc:
                errors.append(f"{label}: フォントの設定に失敗しました: {str(exc)[:80]}")

    border = style.get("border")
    if border is not None:
        if not isinstance(border, dict):
            errors.append(f"{label}: border はオブジェクトで指定してください。")
        else:
            try:
                cur = cell.border

                def _side_spec(side):
                    return {
                        "style": getattr(side, "style", None),
                        "color": getattr(side, "color", None),
                    }

                sides = {
                    "left": _side_spec(cur.left),
                    "right": _side_spec(cur.right),
                    "top": _side_spec(cur.top),
                    "bottom": _side_spec(cur.bottom),
                }
                all_style = border.get("style")
                all_color = border.get("color")
                if all_style is not None:
                    for sname in sides:
                        sides[sname]["style"] = all_style
                if all_color is not None:
                    color = _normalize_xlsx_color(all_color)
                    if color is None:
                        errors.append(f"{label}: 罫線の色が不正です: {all_color}")
                    else:
                        for sname in sides:
                            sides[sname]["color"] = color
                for sname in ("left", "right", "top", "bottom"):
                    side_spec = border.get(sname)
                    if side_spec is None:
                        continue
                    if not isinstance(side_spec, dict):
                        errors.append(f"{label}: border.{sname} はオブジェクトで指定してください。")
                        continue
                    if side_spec.get("style") is not None:
                        sides[sname]["style"] = side_spec["style"]
                    if side_spec.get("color") is not None:
                        color = _normalize_xlsx_color(side_spec["color"])
                        if color is None:
                            errors.append(f"{label}: 罫線の色が不正です: {side_spec['color']}")
                        else:
                            sides[sname]["color"] = color
                new_sides = {}
                for sname, spec in sides.items():
                    if spec["style"] is None and spec["color"] is None:
                        new_sides[sname] = getattr(cur, sname)
                    else:
                        side_kw = {}
                        if spec["style"] is not None:
                            side_kw["style"] = spec["style"]
                        if spec["color"] is not None:
                            side_kw["color"] = spec["color"]
                        new_sides[sname] = Side(**side_kw)
                cell.border = Border(**new_sides)
            except Exception as exc:
                errors.append(f"{label}: 罫線の設定に失敗しました: {str(exc)[:80]}")

    alignment = style.get("alignment")
    if alignment is None:
        alignment = style.get("align")
    if alignment is not None:
        if not isinstance(alignment, dict):
            errors.append(f"{label}: alignment はオブジェクトで指定してください。")
        else:
            try:
                cur = cell.alignment
                kw = {
                    "horizontal": cur.horizontal,
                    "vertical": cur.vertical,
                    "wrap_text": cur.wrap_text,
                    "shrink_to_fit": cur.shrink_to_fit,
                }
                if alignment.get("horizontal") is not None:
                    kw["horizontal"] = str(alignment["horizontal"]).lower()
                if alignment.get("vertical") is not None:
                    kw["vertical"] = str(alignment["vertical"]).lower()
                if alignment.get("wrapText") is not None:
                    kw["wrap_text"] = bool(alignment["wrapText"])
                if alignment.get("shrinkToFit") is not None:
                    kw["shrink_to_fit"] = bool(alignment["shrinkToFit"])
                cell.alignment = Alignment(**kw)
            except Exception as exc:
                errors.append(f"{label}: 配置の設定に失敗しました: {str(exc)[:80]}")

    if style.get("numberFormat") is not None:
        try:
            cell.number_format = str(style["numberFormat"])
        except Exception as exc:
            errors.append(f"{label}: 表示形式の設定に失敗しました: {str(exc)[:80]}")

    return errors


def _apply_xlsx_cell_edits(data, cell_edits, keep_vba=False):
    """Apply targeted cell-value edits to an xlsx/xlsm workbook in place.

    The workbook is opened with openpyxl (not read-only), so all existing
    formatting (fill colors, borders, fonts, merged cells, number formats) is
    preserved and only the given cells' values are changed.  Each edit may also
    carry a ``style`` object (fill / font / border / alignment / numberFormat)
    that adds or updates formatting on that cell while leaving everything else
    untouched.  Returns ``(new_bytes, errors)`` where ``errors`` is a list of
    human-readable messages describing any invalid edits that were skipped.
    """
    from openpyxl import load_workbook

    if not isinstance(cell_edits, list):
        return None, ["cell_edits はリストで指定してください。"]
    try:
        wb = load_workbook(BytesIO(data), keep_vba=bool(keep_vba))
    except Exception as exc:
        raise ValueError(f"Excelファイルを開けませんでした: {str(exc)[:120]}")
    default_sheet = wb.worksheets[0].title if wb.worksheets else None
    errors = []
    for i, edit in enumerate(cell_edits):
        if not isinstance(edit, dict):
            errors.append(f"編集 {i + 1}: 形式が不正です。")
            continue
        cell_ref = str(edit.get("cell") or "").strip()
        if not _XLSX_CELL_RE.match(cell_ref):
            errors.append(f"編集 {i + 1}: セル番地が不正です: {cell_ref or '(空)'}")
            continue
        sheet_name = str(edit.get("sheet") or "").strip() or default_sheet
        if not sheet_name or sheet_name not in wb.sheetnames:
            errors.append(f"編集 {i + 1}: シートが見つかりません: {sheet_name}")
            continue
        label = f"編集 {i + 1} ({sheet_name}!{cell_ref})"
        try:
            cell = wb[sheet_name][cell_ref]
            if "value" in edit:
                cell.value = edit.get("value")
        except Exception as exc:
            errors.append(f"{label}: セルの更新に失敗しました: {str(exc)[:100]}")
            continue
        style = edit.get("style")
        if style is not None:
            errors.extend(_apply_cell_style(cell, style, label))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), errors


# Namespace used by WordprocessingML document.xml.
_WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def _docx_paragraph_text(p):
    """Return the concatenated text of a ``w:p`` lxml element."""
    ns = f"{{{_WORD_NS}}}"
    return "".join(t.text or "" for t in p.findall(f".//{ns}t"))


def _resolve_docx_paragraph(paragraphs, spec):
    """Resolve a paragraph_edits ``paragraph`` spec to a ``w:p`` element.

    ``spec`` may be a 1-based integer index (matching the numbered attachment
    view), a numeric string, or a string to match against the paragraph text
    (exact or partial).  Returns ``None`` when nothing matches.
    """
    if spec is None or isinstance(spec, bool):
        return None
    if isinstance(spec, int):
        if spec < 1 or spec > len(paragraphs):
            return None
        return paragraphs[spec - 1]
    text = str(spec).strip()
    if not text:
        return None
    if text.isdigit():
        idx = int(text)
        if 1 <= idx <= len(paragraphs):
            return paragraphs[idx - 1]
    for p in paragraphs:
        ptxt = _docx_paragraph_text(p)
        if ptxt == text or (text and text in ptxt):
            return p
    return None


def _set_docx_paragraph_text(p, text):
    """Replace a paragraph's text, keeping the first run's formatting."""
    import copy
    ns = f"{{{_WORD_NS}}}"
    xml_space = "{http://www.w3.org/XML/1998/namespace}space"
    all_runs = p.findall(f".//{ns}r")
    rpr_to_keep = None
    if all_runs:
        rpr_to_keep = all_runs[0].find(f"{ns}rPr")
    for run in all_runs:
        parent = run.getparent()
        if parent is not None:
            parent.remove(run)
    run = _LXML.SubElement(p, f"{ns}r")
    if rpr_to_keep is not None:
        run.insert(0, copy.deepcopy(rpr_to_keep))
    t_el = _LXML.SubElement(run, f"{ns}t")
    t_el.text = text
    if text != text.strip():
        t_el.set(xml_space, "preserve")


def _set_docx_onoff(rpr, tag, enabled):
    """Set/update a boolean run property (w:b, w:i, w:strike, ...)."""
    ns = f"{{{_WORD_NS}}}"
    el = rpr.find(f"{ns}{tag}")
    if el is None:
        el = _LXML.SubElement(rpr, f"{ns}{tag}")
    if enabled:
        el.attrib.pop(f"{ns}val", None)
    else:
        el.set(f"{ns}val", "0")
    return el


def _set_docx_val(rpr, tag, value):
    """Set/update a value-carrying run property (w:color, w:sz, w:u, ...)."""
    ns = f"{{{_WORD_NS}}}"
    el = rpr.find(f"{ns}{tag}")
    if el is None:
        el = _LXML.SubElement(rpr, f"{ns}{tag}")
    el.set(f"{ns}val", value)
    return el


def _set_docx_rfonts(rpr, name):
    """Set the font name on a run for all script ranges."""
    ns = f"{{{_WORD_NS}}}"
    el = rpr.find(f"{ns}rFonts")
    if el is None:
        el = _LXML.SubElement(rpr, f"{ns}rFonts")
    for attr in ("ascii", "hAnsi", "eastAsia", "cs"):
        el.set(f"{ns}{attr}", name)
    return el


def _apply_docx_paragraph_style(p, style, label, errors):
    """Apply a style dict to a ``w:p`` element (runs + alignment)."""
    ns = f"{{{_WORD_NS}}}"
    if not isinstance(style, dict):
        errors.append(f"{label}: style はオブジェクトで指定してください。")
        return
    font = style.get("font")
    if font is not None:
        if not isinstance(font, dict):
            errors.append(f"{label}: font はオブジェクトで指定してください。")
        else:
            runs = p.findall(f".//{ns}r")
            if not runs:
                run = _LXML.SubElement(p, f"{ns}r")
                runs = [run]
            for run in runs:
                rpr = run.find(f"{ns}rPr")
                if rpr is None:
                    rpr = _LXML.SubElement(run, f"{ns}rPr")
                    run.insert(0, rpr)
                if font.get("bold") is not None:
                    _set_docx_onoff(rpr, "b", bool(font["bold"]))
                if font.get("italic") is not None:
                    _set_docx_onoff(rpr, "i", bool(font["italic"]))
                if font.get("strikethrough") is not None:
                    _set_docx_onoff(rpr, "strike", bool(font["strikethrough"]))
                if font.get("underline") is not None:
                    u = str(font["underline"]).strip().lower()
                    if u in ("none", "single", "double"):
                        _set_docx_val(rpr, "u", u)
                    else:
                        errors.append(f"{label}: underline の値が不正です: {font['underline']}")
                if font.get("color") is not None:
                    color = _normalize_xlsx_color(font["color"])
                    if color is None:
                        errors.append(f"{label}: 文字色が不正です: {font['color']}")
                    else:
                        # Word uses RRGGBB (drop the ARGB alpha prefix).
                        _set_docx_val(rpr, "color", color[2:])
                if font.get("size") is not None:
                    try:
                        half_points = str(int(round(float(font["size"]) * 2)))
                        _set_docx_val(rpr, "sz", half_points)
                        _set_docx_val(rpr, "szCs", half_points)
                    except (TypeError, ValueError):
                        errors.append(f"{label}: フォントサイズが不正です: {font['size']}")
                if font.get("name") is not None:
                    _set_docx_rfonts(rpr, str(font["name"]))
                if font.get("highlight") is not None:
                    h = str(font["highlight"]).strip().lower()
                    if h in ("none", "yellow", "green", "cyan", "magenta", "red", "blue", "gray"):
                        _set_docx_val(rpr, "highlight", h)
                    else:
                        errors.append(f"{label}: highlight の値が不正です: {font['highlight']}")
    alignment = style.get("alignment")
    if alignment is not None:
        align_val = str(alignment).strip().lower()
        ppr = p.find(f"{ns}pPr")
        if align_val in ("left", "default"):
            if ppr is not None:
                jc = ppr.find(f"{ns}jc")
                if jc is not None:
                    ppr.remove(jc)
        else:
            if ppr is None:
                ppr = _LXML.SubElement(p, f"{ns}pPr")
                p.insert(0, ppr)
            jc = ppr.find(f"{ns}jc")
            if jc is None:
                jc = _LXML.SubElement(ppr, f"{ns}jc")
            jc.set(f"{ns}val", align_val)


def _apply_docx_paragraph_edits(data, paragraph_edits):
    """Apply targeted paragraph edits to a docx in place.

    Each edit may replace the paragraph text and/or add formatting (bold,
    italic, underline, strikethrough, color, size, font name, highlight,
    alignment) while leaving every other property of the document untouched.
    Paragraphs are addressed by 1-based index (the numbered attachment view) or
    by a (partial) text match.  Returns ``(new_bytes, errors)``.
    """
    ns = f"{{{_WORD_NS}}}"
    if not isinstance(paragraph_edits, list):
        return None, ["paragraph_edits はリストで指定してください。"]
    try:
        with zipfile.ZipFile(BytesIO(data)) as zf:
            doc_xml = zf.read('word/document.xml')
            other_entries = [
                (item.filename, zf.read(item.filename))
                for item in zf.infolist()
                if item.filename != 'word/document.xml'
            ]
    except Exception as exc:
        raise ValueError(f"Wordファイルを開けませんでした: {str(exc)[:120]}")
    parser = _LXML.XMLParser(resolve_entities=False, no_network=True)
    try:
        tree = _LXML.fromstring(doc_xml, parser=parser)
    except Exception as exc:
        raise ValueError(f"Wordファイルの内容を解析できませんでした: {str(exc)[:120]}")
    paragraphs = tree.findall(f".//{ns}p")
    if not paragraphs:
        return None, ["段落が見つかりませんでした。"]
    errors = []
    for i, edit in enumerate(paragraph_edits):
        if not isinstance(edit, dict):
            errors.append(f"編集 {i + 1}: 形式が不正です。")
            continue
        spec = edit.get("paragraph")
        p = _resolve_docx_paragraph(paragraphs, spec)
        if p is None:
            errors.append(f"編集 {i + 1}: 段落が見つかりません: {spec or '(空)'}")
            continue
        label = f"編集 {i + 1}"
        if "text" in edit and edit["text"] is not None:
            _set_docx_paragraph_text(p, str(edit["text"]))
        style = edit.get("style")
        if style is not None:
            _apply_docx_paragraph_style(p, style, label, errors)
    try:
        new_doc_xml = _LXML.tostring(
            tree, encoding='UTF-8', xml_declaration=True, standalone=True
        )
    except Exception as exc:
        raise ValueError(f"Wordファイルの保存に失敗しました: {str(exc)[:120]}")
    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('word/document.xml', new_doc_xml)
        for name, content in other_entries:
            zf.writestr(name, content)
    return buf.getvalue(), errors


def _apply_pdf_text_edits(data, text_edits):
    """Apply best-effort in-place text replacements to a PDF.

    PyMuPDF searches for each ``find`` string on the given pages, covers the
    found text with a white rectangle and draws the ``replace`` string at the
    same position, so the surrounding layout is preserved.  This is best-effort:
    text split across lines, images, or text hidden in complex content streams
    cannot be replaced.  Returns ``(new_bytes, errors)``.
    """
    if not isinstance(text_edits, list):
        return None, ["text_edits はリストで指定してください。"]
    if not text_edits:
        return None, ["text_edits を指定してください。"]
    try:
        import pymupdf
    except Exception:
        return None, ["PDFの編集ライブラリが利用できないため、text_edits による編集はできません。"]
    errors = []
    try:
        doc = pymupdf.open(stream=bytes(data), filetype="pdf")
    except Exception as exc:
        raise ValueError(f"PDFファイルを開けませんでした: {str(exc)[:120]}")
    if getattr(doc, "needs_pass", False):
        return None, ["パスワード保護されたPDFのため編集できません。"]
    n_pages = len(doc)
    for i, edit in enumerate(text_edits):
        if not isinstance(edit, dict):
            errors.append(f"編集 {i + 1}: 形式が不正です。")
            continue
        find_text = str(edit.get("find") or "").strip()
        if not find_text:
            errors.append(f"編集 {i + 1}: find（検索する文字列）を指定してください。")
            continue
        replace_text = str(edit.get("replace") or "")
        page_spec = edit.get("page")
        try:
            if page_spec is None:
                page_nos = range(n_pages)
            else:
                pno = int(page_spec)
                if pno < 1 or pno > n_pages:
                    errors.append(f"編集 {i + 1}: ページ番号が範囲外です: {page_spec}")
                    continue
                page_nos = [pno - 1]
            found_any = False
            for pno in page_nos:
                page = doc[pno]
                rects = page.search_for(find_text)
                if not rects:
                    continue
                found_any = True
                for rect in rects:
                    page.add_redact_annot(rect, fill=(1, 1, 1))
                page.apply_redactions()
                for rect in rects:
                    fontsize = max(6.0, min(72.0, rect.height * 0.9))
                    page.insert_text(
                        (rect.x0, rect.y1),
                        replace_text,
                        fontsize=fontsize,
                        fontname="japan",
                        color=(0, 0, 0),
                    )
            if not found_any:
                errors.append(f"編集 {i + 1}: テキストが見つかりませんでした: {find_text}")
        except Exception as exc:
            errors.append(f"編集 {i + 1}: 置換に失敗しました: {str(exc)[:100]}")
    try:
        out = doc.tobytes()
    except Exception as exc:
        raise ValueError(f"PDFの保存に失敗しました: {str(exc)[:120]}")
    return out, errors


def _docx_paragraph_to_html(p):
    """Convert a python-docx Paragraph to an HTML fragment (best-effort)."""
    from docx.enum.text import WD_ALIGN_PARAGRAPH
    from html import escape as _hes

    style_name = p.style.name if p.style else ""
    tag = "p"
    if style_name and style_name.lower().startswith("heading"):
        digits = re.sub(r"[^0-9]", "", style_name)
        if digits:
            tag = f"h{min(int(digits[:1]), 6)}"
    elif style_name and style_name.lower() in ("title",):
        tag = "h1"
    align_css = ""
    if p.alignment in (WD_ALIGN_PARAGRAPH.CENTER, WD_ALIGN_PARAGRAPH.RIGHT, WD_ALIGN_PARAGRAPH.JUSTIFY):
        amap = {
            WD_ALIGN_PARAGRAPH.CENTER: "center",
            WD_ALIGN_PARAGRAPH.RIGHT: "right",
            WD_ALIGN_PARAGRAPH.JUSTIFY: "justify",
        }
        align_css = f"text-align:{amap[p.alignment]};"
    inner = []
    for run in p.runs:
        text = _hes(run.text)
        if not text:
            continue
        open_tags = []
        close_tags = []
        css = []
        if run.bold:
            open_tags.append("<b>")
            close_tags.append("</b>")
        if run.italic:
            open_tags.append("<i>")
            close_tags.append("</i>")
        if run.underline:
            open_tags.append("<u>")
            close_tags.append("</u>")
        if run.font.size:
            css.append(f"font-size:{run.font.size.pt:.1f}pt")
        if run.font.color and run.font.color.rgb:
            css.append(f"color:#{run.font.color.rgb}")
        if run.font.name:
            css.append(f"font-family:'{_hes(run.font.name)}'")
        styled = text
        if css:
            styled = f'<span style="{"; ".join(css)}">{styled}</span>'
        for t_ in open_tags:
            styled = t_ + styled
        for t_ in close_tags:
            styled = styled + t_
        inner.append(styled)
    content = "".join(inner)
    if not content:
        return ""
    attrs = f' style="{align_css}"' if align_css else ""
    return f"<{tag}{attrs}>{content}</{tag}>"


def _docx_table_to_html(table):
    """Convert a python-docx Table to an HTML fragment (best-effort)."""
    rows = []
    for row in table.rows:
        cells = []
        for cell in row.cells:
            cell_html = " ".join(_docx_paragraph_to_html(p) for p in cell.paragraphs)
            cells.append(f"<td>{cell_html}</td>")
        rows.append("<tr>" + "".join(cells) + "</tr>")
    return "<table>" + "".join(rows) + "</table>"


def _convert_docx_to_pdf_bytes(data):
    """Best-effort DOCX -> PDF conversion (python-docx -> HTML -> WeasyPrint).

    Preserves common formatting (headings, bold/italic/underline, colors,
    sizes, alignment, tables) but not pixel-exact Word layout.  Used to provide
    a PDF version alongside an edited Word file.
    """
    from docx import Document
    from docx.oxml.ns import qn

    doc = Document(BytesIO(data))
    parts = []
    for child in doc.element.body.iterchildren():
        if child.tag == qn('w:p'):
            from docx.text.paragraph import Paragraph
            parts.append(_docx_paragraph_to_html(Paragraph(child, doc)))
        elif child.tag == qn('w:tbl'):
            from docx.table import Table
            parts.append(_docx_table_to_html(Table(child, doc)))
    html_body = "\n".join(part for part in parts if part)
    return _render_html_to_pdf_bytes(html_body)


def _render_html_to_pdf_bytes(html_body, title="Document"):
    """Render an HTML fragment to a PDF using WeasyPrint (shared by builders)."""
    import markdown as _md

    safe_title = html.escape(str(title or "Document").strip() or "Document")
    document_html = f"""<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8">
<title>{safe_title}</title>
<style>
@page {{
  size: A4;
  margin: 16mm 16mm 18mm;
  @bottom-right {{
    content: "Page " counter(page) " / " counter(pages);
    color: #6b7280;
    font-size: 8.5pt;
  }}
}}
html {{ color-scheme: light; background: #ffffff; }}
body {{
  margin: 0;
  background: #ffffff;
  color: #1f2937;
  font-family: "IPAPGothic", "IPAGothic", "Droid Sans Fallback", sans-serif;
  font-size: 10.5pt;
  line-height: 1.55;
  overflow-wrap: anywhere;
  word-break: normal;
}}
h1 {{ font-size: 18pt; margin: 0 0 4mm; padding-bottom: 2mm; border-bottom: 1.5pt solid #e5e7eb; break-after: avoid; }}
h2 {{ font-size: 15pt; margin: 5mm 0 2.5mm; break-after: avoid; }}
h3 {{ font-size: 13pt; margin: 4mm 0 2mm; break-after: avoid; }}
h4, h5, h6 {{ font-size: 11pt; margin: 3mm 0 1.5mm; break-after: avoid; }}
p {{ margin: 0 0 0.85em; }}
table {{
  max-width: 100%;
  margin: 1em 0;
  border-collapse: collapse;
  font-size: 9pt;
}}
thead {{ display: table-header-group; }}
tr {{ break-inside: avoid; }}
th, td {{ border: 0.6pt solid #d1d5db; padding: 2.2mm 2.5mm; text-align: left; }}
th {{ background: #f3f4f6; }}
</style>
</head>
<body>
{html_body}
</body>
</html>"""
    weasyprint_bin = os.path.join(os.path.dirname(sys.executable), "weasyprint")
    if not os.path.isfile(weasyprint_bin):
        raise RuntimeError("weasyprint_command_not_found")
    completed = subprocess.run(
        [
            weasyprint_bin,
            "--quiet",
            "--presentational-hints",
            "--optimize-images",
            "--jpeg-quality", "92",
            "--dpi", "180",
            "--timeout", "5",
            "--allowed-protocols", "data",
            "--no-http-redirects",
            "-", "-",
        ],
        input=document_html.encode("utf-8"),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        timeout=90,
        check=False,
    )
    if completed.returncode != 0:
        stderr = completed.stderr.decode("utf-8", errors="replace").strip()
        raise RuntimeError(f"weasyprint_failed: {stderr[-1000:]}")
    pdf_bytes = completed.stdout
    if not isinstance(pdf_bytes, bytes) or not pdf_bytes.startswith(b"%PDF"):
        raise RuntimeError("weasyprint_invalid_pdf")
    return pdf_bytes


def _execute_edit_file_tool(user_id, args, encrypt, loaded_files=None, history=None, thread_id=None):
    """Edit an uploaded or previously created file.

    Editing modes:

    * Excel (xlsx/xlsm): the model supplies ``cell_edits`` (a list of cell
      addresses + new values, optionally with a ``style``).  The original
      workbook is opened in place, so existing formatting is preserved and only
      the given cells change.
    * Word (docx): the model supplies ``paragraph_edits`` (a list of paragraph
      references + new text / ``style``).  The document is edited in place, so
      existing formatting is preserved and new formatting can be added.  A PDF
      version of the edited document is also saved alongside the Word file.
    * PDF: the model supplies ``text_edits`` (a list of find/replace pairs).
      This is best-effort in-place text replacement that keeps the layout.
    * Text-like files (text/markdown/code/csv/tsv/json) and PDF/DOCX fallback:
      the model supplies ``content`` (the full new content; Markdown for
      PDF/DOCX).

    The edited file is saved into the user's library as a new file that keeps
    the original attachment's name and extension, so the original is preserved.

    result = {"ok": True, "filename": ..., "display_name": ..., "url": ...} or
             {"ok": False, "error": ...}
    """
    if not isinstance(args, dict):
        return {"ok": False, "error": "edit_fileの引数が不正です。"}
    source = str(args.get("source") or "").strip()
    if not source:
        return {"ok": False, "error": "source（編集対象の添付ファイル名）は必須です。"}
    target = _resolve_attached_file_ref(
        source,
        loaded_files=loaded_files or [],
        history=history,
        user_id=user_id,
        thread_id=thread_id,
    )
    if not target:
        return {"ok": False, "error": f"添付ファイルが見つかりません: {source}"}
    rel_path = str(target.get('path') or target.get('name') or '').strip()
    if not rel_path:
        return {"ok": False, "error": f"添付ファイルのパスが不明です: {source}"}
    orig_ext = os.path.splitext(rel_path)[1].lower()
    if not _create_file_allowed_ext(orig_ext):
        return {"ok": False, "error": f"対応していないファイル形式です: {orig_ext or '(拡張子なし)'}"}
    format_name = _infer_create_file_format(rel_path)
    # Base the output name on the user-facing attachment name so the edited
    # file keeps the original document's identity (not the random on-disk name).
    # _sanitize_file_display_name preserves non-ASCII (e.g. Japanese) names,
    # unlike the ASCII-only secure_filename path used by create_file.
    display_name = _sanitize_file_display_name(
        target.get('send_name') or os.path.basename(rel_path) or "file"
    ) or "file"
    display_ext = os.path.splitext(display_name)[1]
    if not display_ext:
        display_name = f"{display_name}{orig_ext}"
    elif display_ext.lower() != orig_ext:
        display_name = f"{os.path.splitext(display_name)[0]}{orig_ext}"
    display_stem = os.path.splitext(display_name)[0]
    # On-disk filenames must stay ASCII-safe (secure_filename strips non-ASCII),
    # so derive a sanitized stem from the display name.
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", display_stem).strip("._ ") or "edited"

    edit_note = ""
    extra_files = []
    paragraph_edits = args.get("paragraph_edits")
    text_edits = args.get("text_edits")
    if format_name == "xlsx":
        # Excel: targeted in-place edits that preserve formatting.
        cell_edits = args.get("cell_edits")
        if not isinstance(cell_edits, list) or not cell_edits:
            return {
                "ok": False,
                "error": (
                    "xlsxファイルの編集には cell_edits（変更するセルのリスト）を指定してください。"
                    "content での全体再生成は元の色・罫線・フォントなどの書式が失われるため対応していません。"
                ),
            }
        info = _get_file_disk_info(rel_path)
        original = _load_user_file_bytes(rel_path, info)
        if not original:
            return {"ok": False, "error": "元ファイルの読み込みに失敗しました。"}
        try:
            data, edit_errors = _apply_xlsx_cell_edits(
                original, cell_edits, keep_vba=(orig_ext == ".xlsm")
            )
        except Exception as exc:
            logger.warning("edit_file xlsx apply failed: %s", exc)
            return {"ok": False, "error": f"Excelファイルの編集に失敗しました: {str(exc)[:200]}"}
        if edit_errors:
            edit_note = "一部の編集を適用できませんでした: " + "; ".join(edit_errors[:5])
    elif format_name == "docx" and isinstance(paragraph_edits, list) and paragraph_edits:
        # Word: targeted paragraph edits that preserve the original formatting.
        info = _get_file_disk_info(rel_path)
        original = _load_user_file_bytes(rel_path, info)
        if not original:
            return {"ok": False, "error": "元ファイルの読み込みに失敗しました。"}
        try:
            data, edit_errors = _apply_docx_paragraph_edits(original, paragraph_edits)
        except Exception as exc:
            logger.warning("edit_file docx apply failed: %s", exc)
            return {"ok": False, "error": f"Wordファイルの編集に失敗しました: {str(exc)[:200]}"}
        if edit_errors:
            edit_note = "一部の編集を適用できませんでした: " + "; ".join(edit_errors[:5])
        # Also provide a PDF version of the edited document (best-effort).
        try:
            pdf_data = _convert_docx_to_pdf_bytes(data)
            extra_files.append({
                "ext": ".pdf",
                "bytes": pdf_data,
                "display_name": f"{display_stem}.pdf",
            })
        except Exception as exc:
            logger.warning("edit_file docx->pdf conversion failed: %s", exc)
    elif format_name == "pdf" and isinstance(text_edits, list) and text_edits:
        # PDF: best-effort in-place text replacement that keeps the layout.
        info = _get_file_disk_info(rel_path)
        original = _load_user_file_bytes(rel_path, info)
        if not original:
            return {"ok": False, "error": "元ファイルの読み込みに失敗しました。"}
        try:
            data, edit_errors = _apply_pdf_text_edits(original, text_edits)
        except Exception as exc:
            logger.warning("edit_file pdf apply failed: %s", exc)
            return {"ok": False, "error": f"PDFファイルの編集に失敗しました: {str(exc)[:200]}"}
        if edit_errors:
            edit_note = "一部の編集を適用できませんでした: " + "; ".join(edit_errors[:5])
    else:
        # Text-like files and PDF/DOCX: full content replacement.
        content = args.get("content")
        if content is None:
            return {"ok": False, "error": "content は必須です。"}
        content = str(content)
        if len(content) > _CREATE_FILE_MAX_CONTENT_CHARS:
            return {"ok": False, "error": "ファイル内容が大きすぎます。"}
        try:
            data = _create_file_generated_bytes(display_name, format_name, content, user_id)
        except Exception as exc:
            logger.warning("edit_file generation failed: %s", exc)
            return {"ok": False, "error": f"ファイル生成に失敗しました: {str(exc)[:200]}"}

    if not isinstance(data, (bytes, bytearray)) or not data:
        return {"ok": False, "error": "編集後のファイルが空です。"}
    if len(data) > _CREATE_FILE_MAX_BYTES:
        return {"ok": False, "error": "編集後のファイルが大きすぎます。"}

    def _make_unique_filename():
        return f"{safe_stem}_{int(time.time())}_{os.urandom(3).hex()}{orig_ext}"

    try:
        fname, url = _save_user_generated_bytes_verified(
            user_id, bytes(data), _make_unique_filename, bool(encrypt)
        )
    except StorageLimitError as exc:
        return {"ok": False, "error": str(exc)}
    except Exception as exc:
        logger.warning("edit_file save failed: %s", exc)
        return {"ok": False, "error": f"ライブラリへの保存に失敗しました: {str(exc)[:200]}"}
    result = {"ok": True, "filename": fname, "display_name": display_name, "url": url, "size": len(data)}
    if extra_files:
        saved_extras = []
        for extra in extra_files:
            extra_bytes = extra["bytes"]
            if not isinstance(extra_bytes, (bytes, bytearray)) or not extra_bytes:
                continue
            if len(extra_bytes) > _CREATE_FILE_MAX_BYTES:
                continue
            ext = extra["ext"]

            def _make_extra_unique_filename():
                return f"{safe_stem}_{int(time.time())}_{os.urandom(3).hex()}{ext}"

            try:
                efname, eurl = _save_user_generated_bytes_verified(
                    user_id, bytes(extra_bytes), _make_extra_unique_filename, bool(encrypt)
                )
                saved_extras.append({
                    "filename": efname,
                    "display_name": extra.get("display_name") or f"{display_stem}{ext}",
                    "url": eurl,
                    "size": len(extra_bytes),
                })
            except Exception as exc:
                logger.warning("edit_file extra save failed: %s", exc)
        if saved_extras:
            result["extra_files"] = saved_extras
    if edit_note:
        result["note"] = edit_note
    return result


def _build_edit_file_tool_schema():
    """Return the OpenAI-compatible tool schema for edit_file."""
    return {
        "type": "function",
        "function": {
            "name": "edit_file",
            "description": (
                "会話に添付された既存のファイルを編集し、編集後のファイルをユーザーのファイルライブラリに保存します。"
                "ユーザーが添付したファイルの編集・更新・修正を求めた場合に使用してください（新規作成には create_file を使用）。"
                "source には添付ファイル名（プロンプトの [File: ...] で表示されるファイル名）を指定します。"
                "Excel(xlsx/xlsm)ファイルは cell_edits で変更するセル（セル番地と新しい値）のリストを指定します。"
                "元の色・罫線・フォント・セル結合などの書式はそのまま維持され、指定したセルの値だけが変わります。"
                "各セルの style に書式指定を追加すると、そのセルへ新しい書式（塗りつぶし色・罫線・太字など）を追加できます（指定した項目だけが上書きされます）。"
                "セル番地は添付時に表示される表の列名ヘッダー行（A, B, C, ...）と行番号で指定します（例: B5, AD15）。"
                "Word(docx)ファイルは paragraph_edits で変更する段落（[N] 番号またはテキスト）のリストを指定します。"
                "元の書式はそのまま維持され、各段落の style に新しい書式（太字・色・配置など）を追加できます。編集後は元の Word ファイルと PDF 版の両方が保存されます。"
                "PDFファイルは text_edits で検索・置換のリストを指定すると、レイアウトを保ったままテキストを置き換えます（ベストエフォート。見つからない場合はエラーになります）。"
                "テキスト・Markdown・コード・CSV・TSV・JSON は content に編集後の全文を指定します。"
                "PDF/DOCX の書式を保たない全文置き換えは content に Markdown 形式の全文を指定します。"
                "元ファイルの構造・内容を保ったまま、変更箇所だけを反映してください。"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "source": {
                        "type": "string",
                        "description": "編集対象の添付ファイル名（例: 予定表.xlsx, data.csv）",
                    },
                    "cell_edits": {
                        "type": "array",
                        "description": "Excel(xlsx/xlsm)ファイル用。変更するセルのリスト。",
                        "items": {
                            "type": "object",
                            "properties": {
                                "cell": {
                                    "type": "string",
                                    "description": "セル番地（例: B5, AD15）。列は添付時に表示される列名ヘッダー行（A, B, C, ...）を参照。",
                                },
                                "value": {
                                    "type": "string",
                                    "description": "セルに入れる新しい値。書式のみを変更する場合は省略できます（省略時は現在の値を維持）。",
                                },
                                "sheet": {
                                    "type": "string",
                                    "description": "シート名（省略時は最初のシート）。",
                                },
                                "style": {
                                    "type": "object",
                                    "description": (
                                        "このセルへ追加・更新する新しい書式（省略時は既存の書式を維持）。指定した項目だけが上書きされます。"
                                        "fill: { color: 塗りつぶし色('#FFFF00' / 'FFFF00' / 8桁ARGB), fillType: 'solid'(既定)または 'none'(塗りつぶし解除) }。"
                                        "font: { bold, italic, strikethrough: true/false, size: サイズ(pt), name: フォント名, color: 文字色, underline: 'none'/'single'/'double' }。"
                                        "border: { style: 四辺すべての罫線('none'/'thin'/'medium'/'thick'/'dashed'/'dotted'/'double'), color: 罫線色, left/right/top/bottom: { style, color } で辺ごとに上書き }。"
                                        "alignment: { horizontal: 'left'/'center'/'right'/'justify', vertical: 'top'/'center'/'bottom', wrapText: true/false }。"
                                        "numberFormat: 表示形式（例: '#,##0', '0.00%', 'yyyy/mm/dd'）。"
                                    ),
                                    "properties": {
                                        "fill": {
                                            "type": "object",
                                            "description": "セルの塗りつぶし。",
                                            "properties": {
                                                "color": {"type": "string", "description": "塗りつぶし色。'#RRGGBB' 形式（例: '#FFFF00'）。"},
                                                "fillType": {"type": "string", "enum": ["solid", "none"], "description": "'solid'（既定）で塗りつぶし、'none' で塗りつぶしを解除。"},
                                            },
                                        },
                                        "font": {
                                            "type": "object",
                                            "description": "フォント設定。",
                                            "properties": {
                                                "bold": {"type": "boolean", "description": "太字にするか。true で太字、false で解除。"},
                                                "italic": {"type": "boolean", "description": "斜体にするか。"},
                                                "strikethrough": {"type": "boolean", "description": "取り消し線。"},
                                                "underline": {"type": "string", "enum": ["none", "single", "double"], "description": "下線。'none' で解除。"},
                                                "color": {"type": "string", "description": "文字色。'#RRGGBB' 形式。"},
                                                "size": {"type": "number", "description": "フォントサイズ（pt）。"},
                                                "name": {"type": "string", "description": "フォント名（例: 'Meiryo', 'Arial'）。"},
                                            },
                                        },
                                        "border": {
                                            "type": "object",
                                            "description": "セルの罫線。",
                                            "properties": {
                                                "style": {"type": "string", "enum": ["none", "thin", "medium", "thick", "dashed", "dotted", "double"], "description": "四辺すべてに適用する罫線スタイル。'none' で罫線を解除。"},
                                                "color": {"type": "string", "description": "罫線の色。'#RRGGBB' 形式。"},
                                                "left": {"type": "object", "properties": {"style": {"type": "string", "enum": ["none", "thin", "medium", "thick", "dashed", "dotted", "double"]}, "color": {"type": "string"}}, "description": "左辺のみの設定。"},
                                                "right": {"type": "object", "properties": {"style": {"type": "string", "enum": ["none", "thin", "medium", "thick", "dashed", "dotted", "double"]}, "color": {"type": "string"}}, "description": "右辺のみの設定。"},
                                                "top": {"type": "object", "properties": {"style": {"type": "string", "enum": ["none", "thin", "medium", "thick", "dashed", "dotted", "double"]}, "color": {"type": "string"}}, "description": "上辺のみの設定。"},
                                                "bottom": {"type": "object", "properties": {"style": {"type": "string", "enum": ["none", "thin", "medium", "thick", "dashed", "dotted", "double"]}, "color": {"type": "string"}}, "description": "下辺のみの設定。"},
                                            },
                                        },
                                        "alignment": {
                                            "type": "object",
                                            "description": "セル内の配置。",
                                            "properties": {
                                                "horizontal": {"type": "string", "enum": ["left", "center", "right", "justify", "general"], "description": "水平配置。"},
                                                "vertical": {"type": "string", "enum": ["top", "center", "bottom", "justify", "distributed"], "description": "垂直配置。"},
                                                "wrapText": {"type": "boolean", "description": "セル内で折り返して表示するか。"},
                                            },
                                        },
                                        "numberFormat": {
                                            "type": "string",
                                            "description": "表示形式コード（例: '#,##0', '0.00%', 'yyyy/mm/dd'）。",
                                        },
                                    },
                                    "additionalProperties": False,
                                },
                            },
                            "required": ["cell"],
                        },
                    },
                    "paragraph_edits": {
                        "type": "array",
                        "description": "Word(docx)ファイル用。段落単位で編集するリスト（元の書式を維持したまま編集し、PDF版も生成されます）。",
                        "items": {
                            "type": "object",
                            "properties": {
                                "paragraph": {
                                    "type": "string",
                                    "description": "編集する段落。添付時に表示される [1] [2] ... の番号（1始まり、文字列で指定可）または段落のテキスト（部分一致）。",
                                },
                                "text": {
                                    "type": "string",
                                    "description": "段落の新しいテキスト。省略時は現在のテキストを維持します（書式のみの変更に使用）。",
                                },
                                "style": {
                                    "type": "object",
                                    "description": (
                                        "この段落に追加・更新する新しい書式（省略時は既存の書式を維持）。指定した項目だけが上書きされます。"
                                        "font: { bold, italic, strikethrough: true/false, underline: 'none'/'single'/'double', color: 文字色('#RRGGBB'), size: サイズ(pt), name: フォント名, highlight: 'yellow'/'green'/'cyan'/'magenta'/'red'/'blue'/'gray'/'none' }。"
                                        "alignment: 'left'/'center'/'right'/'justify'/'default'（'default' は既定に戻す）。"
                                    ),
                                    "properties": {
                                        "font": {
                                            "type": "object",
                                            "description": "フォント設定。",
                                            "properties": {
                                                "bold": {"type": "boolean", "description": "太字にするか。true で太字、false で解除。"},
                                                "italic": {"type": "boolean", "description": "斜体にするか。"},
                                                "strikethrough": {"type": "boolean", "description": "取り消し線。"},
                                                "underline": {"type": "string", "enum": ["none", "single", "double"], "description": "下線。'none' で解除。"},
                                                "color": {"type": "string", "description": "文字色。'#RRGGBB' 形式。"},
                                                "size": {"type": "number", "description": "フォントサイズ（pt）。"},
                                                "name": {"type": "string", "description": "フォント名（例: 'Meiryo', 'Arial'）。"},
                                                "highlight": {"type": "string", "enum": ["none", "yellow", "green", "cyan", "magenta", "red", "blue", "gray"], "description": "蛍光ペンの色。'none' で解除。"},
                                            },
                                        },
                                        "alignment": {
                                            "type": "string",
                                            "enum": ["left", "center", "right", "justify", "default"],
                                            "description": "段落の配置。'default' は既定の配置に戻す。",
                                        },
                                    },
                                    "additionalProperties": False,
                                },
                            },
                            "required": ["paragraph"],
                        },
                    },
                    "text_edits": {
                        "type": "array",
                        "description": "PDFファイル用。検索・置換のリスト（ベストエフォート。レイアウトを保ったまま置換します）。",
                        "items": {
                            "type": "object",
                            "properties": {
                                "find": {
                                    "type": "string",
                                    "description": "検索する文字列。",
                                },
                                "replace": {
                                    "type": "string",
                                    "description": "置き換える文字列（空文字で削除）。",
                                },
                                "page": {
                                    "type": "integer",
                                    "description": "置換するページ番号（1始まり）。省略時は全ページ。",
                                },
                            },
                            "required": ["find", "replace"],
                        },
                    },
                    "content": {
                        "type": "string",
                        "description": "テキスト系・PDF・DOCX ファイル用の編集後の新しい内容。",
                    },
                },
                "required": ["source"],
                "additionalProperties": False,
            },
        },
    }

